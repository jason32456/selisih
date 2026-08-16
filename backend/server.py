from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List
import os
import uuid

from io import BytesIO

from dotenv import load_dotenv
from fastapi import APIRouter, FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
import pdfplumber
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")
client = AsyncIOMotorClient(os.environ["MONGO_URL"])
db = client[os.environ["DB_NAME"]]
app = FastAPI(title="SELISIH")
api = APIRouter(prefix="/api")

class ReconciliationCreate(BaseModel):
    supplier_name: str
    reference: str = ""
    po_number: str = ""
    surat_jalan_numbers: List[str] = []
    faktur_number: str = ""
    po_lines: List[Dict[str, Any]]
    delivery_lines: List[Dict[str, Any]]
    invoice_lines: List[Dict[str, Any]]

def money(value: Any) -> int:
    return int(value or 0)

def parse_price(text: str) -> int:
    if not text: return 0
    cleaned = re.sub(r"[Rr][Pp]\.?", "", str(text)).replace(" ", "").strip()
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        parts = cleaned.split(",")
        cleaned = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) <= 2 else cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(".", "")
    try: return int(round(float(cleaned)))
    except Exception: return 0

def parse_pdf_rows(pdf_bytes: bytes, doc_type: str) -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []; failures: List[str] = []
    price_re = re.compile(r"^(?:Rp\.?\s?)?[\d.,]+$", re.IGNORECASE)
    sku_re = re.compile(r"^[A-Z0-9][A-Z0-9\-_/]{2,}$", re.IGNORECASE)
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            tables: List[List[List[str]]] = []
            for page in pdf.pages:
                for tab in page.extract_tables() or []:
                    if tab: tables.append(tab)
            text_lines: List[str] = []
            if not tables:
                for page in pdf.pages:
                    text_lines.extend((page.extract_text() or "").split("\n"))
            for table in tables:
                header_seen = False
                for raw in table:
                    cells = [str(c or "").strip() for c in raw]
                    if not any(cells): continue
                    joined = " ".join(cells).lower()
                    if not header_seen and any(k in joined for k in ("sku", "kode", "nama barang", "produk")):
                        header_seen = True; continue
                    numbers = [parse_price(c) for c in cells if price_re.match(c)]
                    if not cells[0] or len(numbers) < 1:
                        if cells[0]: failures.append(f"Baris tidak terbaca: {' | '.join(cells)}")
                        continue
                    sku = cells[0]; name = cells[1] if len(cells) > 1 else ""
                    unit = cells[2] if len(cells) > 2 and not price_re.match(cells[2]) else "karton"
                    if doc_type == "delivery":
                        rows.append({"sku": sku, "name": name, "unit": unit, "quantity": numbers[0], "damaged": numbers[1] if len(numbers) > 1 else 0})
                    else:
                        rows.append({"sku": sku, "name": name, "unit": unit, "quantity": numbers[0], "unit_price": numbers[-1] if len(numbers) > 1 else 0})
            for line in text_lines:
                line = line.strip()
                if not line: continue
                lower = line.lower()
                if any(k in lower for k in ("purchase order", "surat jalan", "faktur", "no.", "tanggal", "kepada", "dari", "total ", "sub total")):
                    continue
                if any(k in lower for k in ("sku ", "kode ", "nama barang", "harga sat", "jumlah diterima")):
                    continue
                tokens = line.split()
                if len(tokens) < 3 or not sku_re.match(tokens[0]): continue
                sku = tokens[0]
                numeric_positions = [i for i, t in enumerate(tokens) if price_re.match(t) and i > 0]
                if not numeric_positions:
                    failures.append(f"Tidak menemukan angka: {line}"); continue
                nums = [parse_price(tokens[i]) for i in numeric_positions]
                first_num_pos = numeric_positions[0]
                middle = [t for t in tokens[1:first_num_pos] if t.lower() != "rp"]
                unit = "karton"
                if middle and middle[-1].lower() in ("karton", "dus", "pcs", "pack", "botol", "sak"):
                    unit = middle[-1].lower(); middle = middle[:-1]
                name = " ".join(middle).strip() or sku
                if doc_type == "delivery":
                    rows.append({"sku": sku, "name": name, "unit": unit, "quantity": nums[0], "damaged": nums[1] if len(nums) > 1 else 0})
                else:
                    rows.append({"sku": sku, "name": name, "unit": unit, "quantity": nums[0], "unit_price": nums[-1] if len(nums) > 1 else 0})
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(400, f"Gagal membaca PDF: {ex}")
    return {"rows": rows, "failures": failures}

def compare_documents(data: "ReconciliationCreate") -> Dict[str, Any]:
    po = {str(x.get("sku", "")).strip(): x for x in data.po_lines if str(x.get("sku", "")).strip()}
    received: Dict[str, Dict[str, Any]] = {}
    for line in data.delivery_lines:
        sku = str(line.get("sku", "")).strip()
        if not sku: continue
        item = received.setdefault(sku, {"sku": sku, "name": line.get("name", ""), "unit": line.get("unit", ""), "quantity": 0, "damaged": 0})
        item["quantity"] += money(line.get("quantity"))
        item["damaged"] += money(line.get("damaged"))
    invoice = {str(x.get("sku", "")).strip(): x for x in data.invoice_lines if str(x.get("sku", "")).strip()}
    skus = list(dict.fromkeys([*po.keys(), *received.keys(), *invoice.keys()]))
    products = []
    total = 0
    for sku in skus:
        p, d, f = po.get(sku, {}), received.get(sku, {}), invoice.get(sku, {})
        ordered, arrived, billed = money(p.get("quantity")), money(d.get("quantity")), money(f.get("quantity"))
        po_price, billed_price = money(p.get("unit_price")), money(f.get("unit_price"))
        damaged = money(d.get("damaged"))
        issues = []
        if p and arrived < ordered: issues.append({"type": "KURANG KIRIM", "value": (ordered-arrived)*po_price})
        if damaged: issues.append({"type": "BARANG RUSAK", "value": damaged*po_price})
        if f and billed > arrived: issues.append({"type": "TAGIH LEBIH", "value": (billed-arrived)*billed_price})
        if f and p and billed_price != po_price: issues.append({"type": "HARGA TIDAK SESUAI", "value": abs(billed_price-po_price)*billed})
        if f and not p: issues.append({"type": "TIDAK ADA DI PO", "value": billed*billed_price})
        if p and arrived > ordered: issues.append({"type": "KIRIM LEBIH", "value": 0})
        if d and not f: issues.append({"type": "BELUM DITAGIH", "value": 0})
        for issue in issues: total += money(issue["value"])
        products.append({"sku": sku, "name": p.get("name") or f.get("name") or d.get("name", ""), "unit": p.get("unit") or f.get("unit") or d.get("unit", ""), "po": {"quantity": ordered, "unit_price": po_price}, "delivery": {"quantity": arrived, "damaged": damaged}, "invoice": {"quantity": billed, "unit_price": billed_price}, "issues": issues, "impact": sum(money(i["value"]) for i in issues)})
    return {"products": products, "total_discrepancy": total}

def seed_lines(names: List[str], base: str, qty: int, price: int):
    return [{"sku": f"{base}-{i:03d}", "name": n, "unit": "karton", "quantity": qty, "unit_price": price} for i, n in enumerate(names, 1)]

async def ensure_seed():
    if await db.reconciliations.count_documents({}): return
    names = ["Minyak Goreng 2L", "Beras Pulen 5kg", "Gula Pasir", "Kecap Manis", "Mie Instan Goreng", "Susu Kental Manis", "Tepung Terigu", "Teh Celup", "Kopi Bubuk", "Sarden Kaleng", "Biskuit", "Air Mineral"]
    p = seed_lines(names, "MGR-2L", 100, 240000)
    matching = {"supplier_name": "CV Sumber Pangan Jaya", "reference": "REK-2026-0039", "po_number": "PO-0039", "surat_jalan_numbers": ["SJ-0039"], "faktur_number": "FK-0039", "po_lines": p, "delivery_lines": [{**x, "damaged": 0} for x in p], "invoice_lines": p}
    n2 = names + ["Saus Sambal", "Kornet Sapi", "Cokelat Batang", "Susu UHT", "Minuman Teh", "Minyak Goreng 1L"]
    p2 = seed_lines(n2, "SKU-FMCG", 100, 250000)
    d2 = [{**x, "damaged": 0, "quantity": x["quantity"]} for x in p2]; d2[0]["quantity"] = 88; d2[1]["damaged"] = 2; d2[2]["quantity"] = 90; d2[3]["quantity"] = 60
    f2 = [dict(x) for x in p2]; f2[4]["unit_price"] = 280000
    examples = [matching, {"supplier_name": "PT Mitra Boga Nusantara", "reference": "REK-2026-0040", "po_number": "PO-0040", "surat_jalan_numbers": ["SJ-0040-A", "SJ-0040-B"], "faktur_number": "FK-0040", "po_lines": p2, "delivery_lines": d2, "invoice_lines": f2}, {"supplier_name": "UD Anugerah Sentosa", "reference": "REK-2026-0041", "po_number": "PO-0041", "surat_jalan_numbers": ["SJ-0041"], "faktur_number": "FK-0041", "po_lines": p2[:9], "delivery_lines": [{**x, "quantity": 60, "damaged": 0} if x["sku"] == p2[0]["sku"] else {**x, "quantity": x["quantity"], "damaged": 0} for x in p2[:9]], "invoice_lines": [*p2[:9], {"sku": "TAMBAHAN-001", "name": "Barang Tambahan", "unit": "karton", "quantity": 10, "unit_price": 300000}] }]
    for item in examples:
        result = compare_documents(ReconciliationCreate(**item)); item.update(result); item.update({"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat(), "share_token": uuid.uuid4().hex})
        await db.reconciliations.insert_one(item)

@api.get("/")
async def root(): return {"message": "SELISIH siap digunakan"}

@api.get("/reconciliations")
async def list_reconciliations(supplier: str = "", start: str = "", end: str = "", status: str = ""):
    query: Dict[str, Any] = {"deleted_at": {"$exists": False}}
    if supplier: query["supplier_name"] = supplier
    if start or end:
        rng: Dict[str, str] = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end + "T23:59:59"
        query["created_at"] = rng
    if status == "paid": query["status"] = "paid"
    elif status == "open": query["$or"] = [{"status": "open"}, {"status": {"$exists": False}}]
    docs = await db.reconciliations.find(query, {"_id": 0, "po_lines": 0, "delivery_lines": 0, "invoice_lines": 0, "products": 0, "history": 0}).sort("created_at", -1).to_list(1000)
    now = datetime.now(timezone.utc)
    for doc in docs:
        created = doc.get("created_at")
        if created:
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                doc["days_open"] = max(0, (now - dt).days)
            except Exception:
                doc["days_open"] = 0
        else:
            doc["days_open"] = 0
    all_docs = await db.reconciliations.find({"deleted_at": {"$exists": False}}, {"_id": 0, "supplier_name": 1, "created_at": 1, "total_discrepancy": 1, "status": 1}).to_list(5000)
    supplier_options = sorted({d.get("supplier_name") for d in all_docs if d.get("supplier_name")})
    suppliers: Dict[str, Dict[str, Any]] = {}
    for doc in docs:
        if doc.get("status") == "paid": continue
        name = doc.get("supplier_name") or "-"
        entry = suppliers.setdefault(name, {"supplier_name": name, "reconciliation_count": 0, "total_discrepancy": 0, "last_activity": doc.get("created_at")})
        entry["reconciliation_count"] += 1
        entry["total_discrepancy"] += money(doc.get("total_discrepancy"))
        if doc.get("created_at") and doc["created_at"] > entry["last_activity"]:
            entry["last_activity"] = doc["created_at"]
    ledger = sorted(suppliers.values(), key=lambda x: x["total_discrepancy"], reverse=True)
    total_open = sum(money(x.get("total_discrepancy")) for x in docs if x.get("status") != "paid")
    this_month = now.strftime("%Y-%m"); last_month = (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    def in_month(d, key): return (d.get("created_at") or "").startswith(key) and d.get("status") != "paid"
    monthly = {"current_month": sum(money(d.get("total_discrepancy")) for d in all_docs if in_month(d, this_month)), "previous_month": sum(money(d.get("total_discrepancy")) for d in all_docs if in_month(d, last_month)), "current_count": sum(1 for d in all_docs if in_month(d, this_month)), "previous_count": sum(1 for d in all_docs if in_month(d, last_month)), "current_label": now.strftime("%B %Y"), "previous_label": (now.replace(day=1) - timedelta(days=1)).strftime("%B %Y")}
    weeks: List[Dict[str, Any]] = []
    today = now.date(); monday = today - timedelta(days=today.weekday())
    for i in range(7, -1, -1):
        start_dt = monday - timedelta(weeks=i); end_dt = start_dt + timedelta(days=7)
        total = sum(money(d.get("total_discrepancy")) for d in all_docs if d.get("status") != "paid" and d.get("created_at") and start_dt.isoformat() <= d["created_at"][:10] < end_dt.isoformat())
        count = sum(1 for d in all_docs if d.get("status") != "paid" and d.get("created_at") and start_dt.isoformat() <= d["created_at"][:10] < end_dt.isoformat())
        weeks.append({"week_start": start_dt.isoformat(), "label": start_dt.strftime("%d %b"), "total": total, "count": count})
    return {"items": docs, "total_found": total_open, "supplier_ledger": ledger, "supplier_options": supplier_options, "monthly": monthly, "weekly_trend": weeks}

@api.get("/reconciliations/{rid}")
async def get_reconciliation(rid: str):
    doc = await db.reconciliations.find_one({"$or": [{"id": rid}, {"share_token": rid}], "deleted_at": {"$exists": False}}, {"_id": 0})
    if not doc: raise HTTPException(404, "Rekonsiliasi tidak ditemukan")
    return doc

@api.delete("/reconciliations/{rid}")
async def delete_reconciliation(rid: str):
    result = await db.reconciliations.update_one({"id": rid, "deleted_at": {"$exists": False}}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    if not result.modified_count: raise HTTPException(404, "Rekonsiliasi tidak ditemukan")
    return {"message": "Rekonsiliasi dihapus", "id": rid}

@api.post("/reconciliations/{rid}/restore")
async def restore_reconciliation(rid: str):
    result = await db.reconciliations.update_one({"id": rid}, {"$unset": {"deleted_at": ""}})
    if not result.modified_count: raise HTTPException(404, "Rekonsiliasi tidak dapat dipulihkan")
    return {"message": "Rekonsiliasi dipulihkan"}

class StatusUpdate(BaseModel):
    status: str
    actor: str = ""

class NotesUpdate(BaseModel):
    notes: str
    actor: str = ""

class BulkAction(BaseModel):
    ids: List[str]
    action: str
    actor: str = ""

def record_history(action: str, actor: str, detail: str = "") -> Dict[str, Any]:
    return {"action": action, "actor": actor or "Pengguna", "detail": detail, "timestamp": datetime.now(timezone.utc).isoformat()}

@api.patch("/reconciliations/{rid}/status")
async def update_status(rid: str, payload: StatusUpdate):
    if payload.status not in ("open", "paid"): raise HTTPException(400, "Status tidak valid")
    entry = record_history("status", payload.actor, f"Status diubah menjadi {payload.status}")
    result = await db.reconciliations.update_one({"id": rid}, {"$set": {"status": payload.status, "settled_at": datetime.now(timezone.utc).isoformat() if payload.status == "paid" else None}, "$push": {"history": entry}})
    if not result.matched_count: raise HTTPException(404, "Rekonsiliasi tidak ditemukan")
    return {"message": "Status diperbarui", "status": payload.status}

@api.patch("/reconciliations/{rid}/notes")
async def update_notes(rid: str, payload: NotesUpdate):
    entry = record_history("notes", payload.actor, "Catatan diperbarui")
    result = await db.reconciliations.update_one({"id": rid, "deleted_at": {"$exists": False}}, {"$set": {"notes": payload.notes[:2000]}, "$push": {"history": entry}})
    if not result.matched_count: raise HTTPException(404, "Rekonsiliasi tidak ditemukan")
    return {"message": "Catatan disimpan"}

@api.post("/reconciliations/bulk")
async def bulk_action(payload: BulkAction):
    if not payload.ids: raise HTTPException(400, "Tidak ada rekonsiliasi terpilih")
    now_iso = datetime.now(timezone.utc).isoformat()
    if payload.action == "delete":
        entry = record_history("delete", payload.actor, "Dihapus (bulk)")
        result = await db.reconciliations.update_many({"id": {"$in": payload.ids}, "deleted_at": {"$exists": False}}, {"$set": {"deleted_at": now_iso}, "$push": {"history": entry}})
    elif payload.action in ("paid", "open"):
        entry = record_history("status", payload.actor, f"Status diubah menjadi {payload.action} (bulk)")
        result = await db.reconciliations.update_many({"id": {"$in": payload.ids}}, {"$set": {"status": payload.action, "settled_at": now_iso if payload.action == "paid" else None}, "$push": {"history": entry}})
    else:
        raise HTTPException(400, "Aksi tidak valid")
    return {"message": "Berhasil", "affected": result.modified_count}

@api.post("/parse-pdf")
async def parse_pdf(doc_type: str, file: UploadFile = File(...)):
    if doc_type not in ("po", "delivery", "faktur"): raise HTTPException(400, "Jenis dokumen tidak valid")
    if not file.filename or not file.filename.lower().endswith(".pdf"): raise HTTPException(400, "Unggah berkas PDF")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024: raise HTTPException(400, "Ukuran PDF terlalu besar (maksimal 5 MB)")
    result = parse_pdf_rows(data, doc_type)
    return result

@api.get("/summary.xlsx")
async def download_summary_xlsx(supplier: str = "", start: str = "", end: str = ""):
    query: Dict[str, Any] = {"deleted_at": {"$exists": False}}
    if supplier: query["supplier_name"] = supplier
    if start or end:
        rng: Dict[str, str] = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end + "T23:59:59"
        query["created_at"] = rng
    docs = await db.reconciliations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    wb = Workbook(); ws = wb.active; ws.title = "Ringkasan"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF"); header_fill = PatternFill("solid", fgColor="14161C")
    ws["A1"] = "RINGKASAN REKONSILIASI"; ws["A1"].font = Font(bold=True, size=16)
    filters = []
    if supplier: filters.append(f"Pemasok: {supplier}")
    if start: filters.append(f"Dari: {start}")
    if end: filters.append(f"Sampai: {end}")
    ws["A2"] = " · ".join(filters) if filters else "Semua rekonsiliasi"
    ws["A3"] = f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    headers = ["Referensi", "Pemasok", "Tanggal", "Status", "Total selisih (Rp)", "Catatan"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=value); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="left")
    row = 6; open_total = 0; paid_total = 0
    for doc in docs:
        status = doc.get("status", "open"); disc = money(doc.get("total_discrepancy"))
        ws.cell(row=row, column=1, value=doc.get("reference", ""))
        ws.cell(row=row, column=2, value=doc.get("supplier_name", ""))
        ws.cell(row=row, column=3, value=(doc.get("created_at") or "")[:10])
        ws.cell(row=row, column=4, value="Sudah dibayar" if status == "paid" else "Terbuka")
        ws.cell(row=row, column=5, value=disc).number_format = "#,##0"
        ws.cell(row=row, column=6, value=doc.get("notes", ""))
        if status == "paid": paid_total += disc
        else: open_total += disc
        row += 1
    ws.cell(row=row + 1, column=4, value="Total terbuka").font = Font(bold=True)
    ws.cell(row=row + 1, column=5, value=open_total).number_format = "#,##0"
    ws.cell(row=row + 2, column=4, value="Total sudah dibayar").font = Font(bold=True)
    ws.cell(row=row + 2, column=5, value=paid_total).number_format = "#,##0"
    for col, width in enumerate([16, 30, 14, 18, 20, 45], 1):
        ws.column_dimensions[chr(64 + col)].width = width
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    filename = f"ringkasan-rekonsiliasi-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api.get("/reconciliations/{rid}/nota.xlsx")
async def download_nota_xlsx(rid: str):
    doc = await db.reconciliations.find_one({"id": rid}, {"_id": 0})
    if not doc: raise HTTPException(404, "Rekonsiliasi tidak ditemukan")
    wb = Workbook(); ws = wb.active; ws.title = "Nota Retur"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="14161C")
    ws["A1"] = "NOTA RETUR"; ws["A1"].font = Font(name="Calibri", bold=True, size=16)
    ws["A2"] = f"Nomor nota: NR-{doc.get('reference', '')}"
    ws["A3"] = f"Pemasok: {doc.get('supplier_name', '')}"
    ws["A4"] = f"Referensi: {doc.get('reference', '')}"
    ws["A5"] = f"Tanggal: {datetime.now().strftime('%d/%m/%Y')}"
    headers = ["SKU", "Produk", "Jenis selisih", "Nilai (Rp)"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=value); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="left")
    row = 8
    for product in doc.get("products", []):
        for issue in product.get("issues", []):
            if not money(issue.get("value")): continue
            ws.cell(row=row, column=1, value=product.get("sku", ""))
            ws.cell(row=row, column=2, value=product.get("name", ""))
            ws.cell(row=row, column=3, value=issue.get("type", ""))
            ws.cell(row=row, column=4, value=money(issue.get("value"))).number_format = "#,##0"
            row += 1
    ws.cell(row=row + 1, column=3, value="Total klaim").font = Font(bold=True)
    total_cell = ws.cell(row=row + 1, column=4, value=money(doc.get("total_discrepancy")))
    total_cell.number_format = "#,##0"; total_cell.font = Font(bold=True)
    for col, width in enumerate([16, 32, 22, 18], 1):
        ws.column_dimensions[chr(64 + col)].width = width
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    filename = f"nota-retur-{doc.get('reference', rid)}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@api.post("/reconciliations")
async def create_reconciliation(data: ReconciliationCreate):
    result = compare_documents(data); doc = data.model_dump(); doc.update(result); doc.update({"id": str(uuid.uuid4()), "reference": data.reference or f"REK-{datetime.now().year}-{uuid.uuid4().hex[:4].upper()}", "created_at": datetime.now(timezone.utc).isoformat(), "share_token": uuid.uuid4().hex})
    await db.reconciliations.insert_one(doc)
    return {key: value for key, value in doc.items() if key != "_id"}

app.include_router(api)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","), allow_methods=["*"], allow_headers=["*"])
@app.on_event("startup")
async def startup(): await ensure_seed()
@app.on_event("shutdown")
async def shutdown(): client.close()